from __future__ import annotations

import io
import re
from collections import defaultdict
from typing import Any

from openpyxl import load_workbook

from crewai_enterprise_pipeline_api.domain.models import (
    FinancialPeriod,
    FinancialStatement,
    QoEAdjustment,
)

PERIOD_PATTERN = re.compile(
    r"(?:(q(?P<quarter>[1-4]))\s*)?(?:fy\s*'?(?P<fy>\d{2,4})|(?P<year>20\d{2}|19\d{2}))",
    re.IGNORECASE,
)

ADJUSTMENT_KEYWORDS = (
    "one-time",
    "one time",
    "non-recurring",
    "non recurring",
    "exceptional",
    "extraordinary",
    "legal cost",
    "restructuring",
    "settlement",
)

HEADER_KEYWORDS = {"particulars", "line item", "metric", "account"}
PERCENT_FIELDS = {"customer_concentration_top_3", "q4_revenue_share"}

METRIC_ALIASES: dict[str, tuple[str, ...]] = {
    "revenue": (
        "revenue",
        "revenue from operations",
        "operating revenue",
        "total income",
        "sales",
    ),
    "gross_profit": ("gross profit",),
    "operating_profit": ("operating profit", "ebit", "profit from operations"),
    "ebitda": ("ebitda", "reported ebitda", "adjusted ebitda"),
    "pat": (
        "pat",
        "profit after tax",
        "net income",
        "net profit",
        "profit for the year",
    ),
    "operating_cash_flow": (
        "operating cash flow",
        "cash flow from operations",
        "net cash from operating activities",
    ),
    "net_debt": ("net debt", "total debt", "debt"),
    "interest_expense": ("interest expense", "finance cost", "finance costs"),
    "working_capital": ("working capital", "net working capital"),
    "total_assets": ("total assets",),
    "shareholder_equity": (
        "shareholder equity",
        "shareholders equity",
        "total equity",
        "net worth",
    ),
    "customer_concentration_top_3": (
        "top 3 customer concentration",
        "top three customer concentration",
        "top 3 customers",
        "top three customers",
    ),
    "q4_revenue_share": (
        "q4 revenue share",
        "q4 revenue % of annual",
        "q4 revenue percentage",
        "q4 annual revenue share",
    ),
}


def _clean_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_metric_label(label: str) -> str | None:
    lowered = _clean_label(label).lower()
    if not lowered:
        return None

    best_match: tuple[int, int, str] | None = None
    for field_name, aliases in METRIC_ALIASES.items():
        for alias in aliases:
            normalized_alias = _clean_label(alias).lower()
            if lowered == normalized_alias:
                candidate = (2, len(normalized_alias), field_name)
            elif re.search(
                rf"(?<![a-z0-9]){re.escape(normalized_alias)}(?![a-z0-9])",
                lowered,
            ):
                candidate = (1, len(normalized_alias), field_name)
            else:
                continue
            if best_match is None or candidate > best_match:
                best_match = candidate
    return best_match[2] if best_match is not None else None


def _period_sort_key(label: str) -> tuple[int, int]:
    match = PERIOD_PATTERN.search(label)
    if match is None:
        return (9999, 99)

    year_text = match.group("year") or match.group("fy") or "9999"
    year = int(year_text)
    if year < 100:
        year += 2000
    quarter_text = match.group("quarter")
    quarter = int(quarter_text) if quarter_text else 5
    return (year, quarter)


def _normalize_period_label(value: Any) -> str | None:
    label = _clean_label(value)
    if not label:
        return None

    match = PERIOD_PATTERN.search(label)
    if match is None:
        return None

    year_text = match.group("year") or match.group("fy")
    if year_text is None:
        return None
    year = int(year_text)
    if year < 100:
        year += 2000

    quarter_text = match.group("quarter")
    if quarter_text:
        return f"Q{quarter_text} FY{str(year)[-2:]}"
    return f"FY{str(year)[-2:]}"


def _parse_numeric(value: Any, *, field_name: str | None = None) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = _clean_label(value)
        if not text:
            return None

        negative = False
        if text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1]

        lowered = text.lower()
        multiplier = 1.0
        if "crore" in lowered or lowered.endswith("cr"):
            multiplier = 10_000_000.0
        elif "lakh" in lowered or lowered.endswith("lac"):
            multiplier = 100_000.0
        elif "million" in lowered or lowered.endswith("mn"):
            multiplier = 1_000_000.0
        elif "billion" in lowered or lowered.endswith("bn"):
            multiplier = 1_000_000_000.0

        text = re.sub(r"[^0-9.+%-]", "", text)
        if text in {"", "-", ".", "-.", "%"}:
            return None
        percent = "%" in text or field_name in PERCENT_FIELDS
        text = text.replace("%", "")
        try:
            number = float(text) * multiplier
        except ValueError:
            return None
        if negative:
            number = -number
        if percent:
            if abs(number) > 1:
                number = number / 100.0
    return number


def _dedupe_adjustments(adjustments: list[QoEAdjustment]) -> list[QoEAdjustment]:
    unique: dict[tuple[str, str], QoEAdjustment] = {}
    for adjustment in adjustments:
        unique[(adjustment.label.lower(), adjustment.category.lower())] = adjustment
    return sorted(unique.values(), key=lambda item: (item.category, item.label))


def _rows_from_markdown(text: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line.startswith("|"):
            continue
        parts = [cell.strip() for cell in line.strip("|").split("|")]
        if not parts or all(set(cell) <= {"-"} for cell in parts):
            continue
        rows.append(parts)
    return rows


class FinancialParser:
    def parse_document(
        self,
        *,
        artifact_id: str | None,
        artifact_title: str,
        document_kind: str,
        filename: str,
        mime_type: str | None,
        parser_name: str | None,
        content: bytes,
        fallback_text: str | None = None,
    ) -> FinancialStatement | None:
        lower_name = filename.lower()
        periods: list[FinancialPeriod] = []
        adjustments: list[QoEAdjustment] = []

        if lower_name.endswith((".xlsx", ".xlsm")):
            periods, adjustments = self._parse_xlsx(content)

        if not periods and fallback_text:
            periods, text_adjustments = self._parse_markdown_like_text(fallback_text)
            adjustments.extend(text_adjustments)

        if not periods and not adjustments:
            return None

        flags = self._source_level_flags(periods)
        return FinancialStatement(
            artifact_id=artifact_id,
            artifact_title=artifact_title,
            document_kind=document_kind,
            parser_name=parser_name or "financial-parser",
            periods=periods,
            qoe_adjustments=_dedupe_adjustments(adjustments),
            flags=flags,
        )

    def _parse_xlsx(self, content: bytes) -> tuple[list[FinancialPeriod], list[QoEAdjustment]]:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        period_map: dict[str, dict[str, float | None]] = defaultdict(dict)
        adjustments: list[QoEAdjustment] = []

        for worksheet in workbook.worksheets[:10]:
            rows = [
                [cell for cell in row]
                for row in worksheet.iter_rows(values_only=True, max_row=400, max_col=16)
            ]
            self._extract_horizontal_grid(rows, period_map, adjustments)
            self._extract_vertical_grid(rows, period_map, adjustments)

        periods = self._periods_from_map(period_map)
        return periods, adjustments

    def _parse_markdown_like_text(
        self,
        text: str,
    ) -> tuple[list[FinancialPeriod], list[QoEAdjustment]]:
        rows = _rows_from_markdown(text)
        period_map: dict[str, dict[str, float | None]] = defaultdict(dict)
        adjustments: list[QoEAdjustment] = []

        if rows:
            self._extract_horizontal_grid(rows, period_map, adjustments)
            self._extract_vertical_grid(rows, period_map, adjustments)

        periods = self._periods_from_map(period_map)
        return periods, adjustments

    def _extract_horizontal_grid(
        self,
        rows: list[list[Any]],
        period_map: dict[str, dict[str, float | None]],
        adjustments: list[QoEAdjustment],
    ) -> None:
        for row_index, row in enumerate(rows[:60]):
            normalized_periods = [_normalize_period_label(cell) for cell in row]
            period_columns = [
                (index, label)
                for index, label in enumerate(normalized_periods)
                if label is not None and index > 0
            ]
            if len(period_columns) < 2:
                continue
            if not self._looks_like_header_row(row):
                continue

            for data_row in rows[row_index + 1 : row_index + 120]:
                if not data_row:
                    continue
                label = _clean_label(data_row[0]) if data_row else ""
                if not label:
                    continue

                metric = _normalize_metric_label(label)
                if metric is not None:
                    for column_index, period_label in period_columns:
                        if column_index >= len(data_row):
                            continue
                        value = _parse_numeric(data_row[column_index], field_name=metric)
                        if value is None:
                            continue
                        period_map[period_label][metric] = value
                    continue

                if not self._is_adjustment_label(label):
                    continue

                latest_value = None
                for column_index, _ in reversed(period_columns):
                    if column_index >= len(data_row):
                        continue
                    latest_value = _parse_numeric(data_row[column_index])
                    if latest_value is not None:
                        break
                if latest_value is None:
                    continue
                adjustments.append(
                    QoEAdjustment(
                        label=label,
                        amount=latest_value,
                        category=self._categorize_adjustment(label),
                    )
                )

    def _extract_vertical_grid(
        self,
        rows: list[list[Any]],
        period_map: dict[str, dict[str, float | None]],
        adjustments: list[QoEAdjustment],
    ) -> None:
        if not rows:
            return
        header = [_normalize_metric_label(cell) for cell in rows[0]]
        metric_columns = [
            (index, field_name)
            for index, field_name in enumerate(header)
            if field_name is not None and index > 0
        ]
        if len(metric_columns) < 2:
            return

        for row in rows[1:120]:
            if not row:
                continue
            period_label = _normalize_period_label(row[0])
            if period_label is None:
                continue
            for column_index, field_name in metric_columns:
                if column_index >= len(row):
                    continue
                value = _parse_numeric(row[column_index], field_name=field_name)
                if value is None:
                    continue
                period_map[period_label][field_name] = value

        if header[0] and self._is_adjustment_label(_clean_label(rows[0][0])):
            for row in rows[1:]:
                label = _clean_label(row[0])
                if not label or not self._is_adjustment_label(label):
                    continue
                latest_value = next(
                    (
                        _parse_numeric(row[column_index])
                        for column_index, _ in reversed(metric_columns)
                        if column_index < len(row)
                        and _parse_numeric(row[column_index]) is not None
                    ),
                    None,
                )
                if latest_value is None:
                    continue
                adjustments.append(
                    QoEAdjustment(
                        label=label,
                        amount=latest_value,
                        category=self._categorize_adjustment(label),
                    )
                )

    def _periods_from_map(
        self,
        period_map: dict[str, dict[str, float | None]],
    ) -> list[FinancialPeriod]:
        periods = [
            FinancialPeriod(label=label, **values)
            for label, values in sorted(
                period_map.items(),
                key=lambda item: _period_sort_key(item[0]),
            )
            if values and label.startswith("FY")
        ]
        return periods

    def _looks_like_header_row(self, row: list[Any]) -> bool:
        if not row:
            return False
        first_cell = _clean_label(row[0]).lower()
        if first_cell in HEADER_KEYWORDS:
            return True
        return bool(first_cell) and "particular" in first_cell

    def _is_adjustment_label(self, label: str) -> bool:
        lowered = label.lower()
        return any(keyword in lowered for keyword in ADJUSTMENT_KEYWORDS)

    def _categorize_adjustment(self, label: str) -> str:
        lowered = label.lower()
        if "legal" in lowered:
            return "legal"
        if "restruct" in lowered:
            return "restructuring"
        if "settlement" in lowered:
            return "settlement"
        if "exceptional" in lowered or "extraordinary" in lowered:
            return "exceptional"
        return "non_recurring"

    def _source_level_flags(self, periods: list[FinancialPeriod]) -> list[str]:
        flags: list[str] = []
        if not periods:
            return flags

        latest = periods[-1]
        if latest.customer_concentration_top_3 and latest.customer_concentration_top_3 > 0.60:
            flags.append(
                "Revenue concentration: top 3 customers exceed 60% of the latest period."
            )
        if (
            latest.operating_cash_flow is not None
            and latest.operating_cash_flow < 0
            and latest.ebitda is not None
            and latest.ebitda > 0
        ):
            flags.append(
                "Cash conversion concern: operating cash flow is negative despite positive EBITDA."
            )
        if latest.q4_revenue_share and latest.q4_revenue_share > 0.40:
            flags.append(
                "Revenue seasonality concern: Q4 contributes more than 40% of annual revenue."
            )
        return flags
